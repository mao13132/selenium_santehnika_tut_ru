# ---------------------------------------------
# Program by @developer_telegrams
#
#
# Version   Date        Info
# 1.0       2023    Initial Version
#
# ---------------------------------------------
import json
import os
import random

from openpyxl.workbook import Workbook
from openpyxl.styles import Font

from settings import patch_project
from src.business.save_exel.formate_all_specifications import formate_all_specifications


class StartSave:
    def __init__(self, settings):
        self.settings = settings

        self.colums_checker = {}

        self.BotDB = settings['BotDB']

        self.colums = ['ID продукта', 'Код', 'Имя продукта', 'Цена', 'Старая цена', 'Комплект главные',
                       'Комплект дополнительные', 'Фото', 'Категория', 'Производитель', 'Тип товара',
                       'Ссылка на сторонний сайт', 'Алгоритм', 'Ед.Измерения', 'ID', 'PARENT_ID', 'Видимость',
                       'Видимость варианта', 'Статус товара', 'Количество', 'Описание', 'Видео', 'Документы',
                       'Гарантия', 'Артикул', 'Страна', 'Производитель', 'Коллекция', 'Цвет']

    def create_number_uncolums(self, ws):

        global_count = 0
        start_count = 33

        for col in range(len(self.colums_harakt) + 5):
            ws.cell(row=1, column=start_count + global_count).value = random.randint(1111, 9999)

            global_count += 1

        return self.colums_checker

    def create_title(self, ws):

        global_count = 0

        for count, col in enumerate(self.colums):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            global_count += 1

        for count, col in enumerate(self.colums_harakt):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            self.colums_checker[col] = global_count + 1

            global_count += 1

        return self.colums_checker

    def write_data(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = ''
        ws.cell(row=count_def, column=2).value = ''
        try:
            # name = post['full_name']
            name = post[2]
        except:
            name = ''
        ws.cell(row=count_def, column=3).value = name

        try:
            # price = int(post['price'])
            price = int(post[3])
        except:
            price = ''

        ws.cell(row=count_def, column=4).value = price
        ws.cell(row=count_def, column=5).value = ''
        ws.cell(row=count_def, column=6).value = ''
        ws.cell(row=count_def, column=7).value = ''

        try:
            image = post[9]
            # image = post['image']
        except:
            image = ''

        ws.cell(row=count_def, column=8).value = image

        try:
            category = post[10]
        except:
            category = ''

        ws.cell(row=count_def, column=9).value = category
        try:
            proiz = post[4]
            # proiz = post['prozvoditel']
        except:
            proiz = ''
        ws.cell(row=count_def, column=10).value = proiz
        ws.cell(row=count_def, column=11).value = ''
        try:
            link = post[1]
            # link = post['link']
        except:
            link = ''
        ws.cell(row=count_def, column=12).value = link
        ws.cell(row=count_def, column=13).value = ''
        try:
            ed = 'шт'
            # ed = post['edinicha']
        except:
            ed = ''
        ws.cell(row=count_def, column=14).value = ed
        ws.cell(row=count_def, column=15).value = ''
        ws.cell(row=count_def, column=16).value = ''
        ws.cell(row=count_def, column=17).value = 1
        ws.cell(row=count_def, column=18).value = 1
        ws.cell(row=count_def, column=19).value = ''
        ws.cell(row=count_def, column=20).value = 9999
        try:
            opis = post[8]
            # opis = post['text']
        except:
            opis = ''
        ws.cell(row=count_def, column=21).value = opis
        ws.cell(row=count_def, column=22).value = ''
        ws.cell(row=count_def, column=23).value = post[12]
        ws.cell(row=count_def, column=24).value = post[13]
        try:
            artikl = post[5]
            # artikl = post['artikle']
        except:
            artikl = ''
        ws.cell(row=count_def, column=25).value = artikl
        try:
            # country = post['xarakt']['Страна']
            country = post[6]
        except:
            country = ''
        ws.cell(row=count_def, column=26).value = country
        ws.cell(row=count_def, column=27).value = proiz
        try:
            coll_name = post['collection']
        except:
            coll_name = ''
        ws.cell(row=count_def, column=28).value = coll_name
        ws.cell(row=count_def, column=29).value = post[7]
        try:
            size = post['xarakt']['Размер']
        except:
            size = ''

        ws.cell(row=count_def, column=30).value = size

        try:
            naznach = post['xarakt']['Помещение']
        except:
            naznach = ''
        ws.cell(row=count_def, column=31).value = naznach
        try:
            sostav_collection = post['xarakt']['Назначение']
        except:
            sostav_collection = ''
        ws.cell(row=count_def, column=32).value = sostav_collection

        count = 0
        start_count = 33

        specifications = json.loads(post[11])

        # for key, value in harakter.items():
        for spec in specifications:
            key = spec['name']

            value = spec['value']

            try:
                ws.cell(row=count_def, column=self.colums_checker[key]).value = value
            except:
                continue
            # ws.cell(row=count_def + count, column=start_count).value = comment['author_comment']

            count += 1
            start_count += 1

        return True

    def itter_rows(self, ws):
        count_def = 3
        for count_post, post in enumerate(self.good_dict):

            try:
                write_data = self.write_data(ws, count_def, post)
            except Exception as es:
                print(f'SaveResult: Исключение {es}')

            count_def += 1

        return True

    def one_sheet(self, ws):

        self.create_number_uncolums(ws)

        name_colums_dict = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def start_save(self, filename):
        self.good_dict = self.BotDB.get_all_products()

        if not self.good_dict:
            return False

        wb = Workbook()

        ws = wb.active

        self.colums_harakt = formate_all_specifications(self.good_dict)

        result = self.one_sheet(ws)

        save_file_name = patch_project + os.sep + filename + '.xlsx'

        wb.save(save_file_name)

        return save_file_name


if __name__ == '__main__':
    from src.sql.bot_connector import BotDB

    settings_save = {
        'BotDB': BotDB
    }

    res_save = StartSave(settings_save).start_save('test')

    print()
